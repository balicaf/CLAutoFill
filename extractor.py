#pip install openpyxl
#pip install Pillow
#pip install pywin32
#pip install --upgrade easygui
#pip install python-can
#pip install Pillow
#pip install pyinstaller

#to make the exe
#cd D:\CLAutoFill-master
#pyinstaller --onefile extractor.py
from PIL import Image
import pythoncom
import openpyxl as O
import time
import win32com.client as comclt
import easygui as eg
import os
import can

oShell= comclt.Dispatch("WScript.Shell")

title1 = "Raw_values.vbs Done?"
title2 = "Step 2"
questionRawValueDone = "Have you already done the Raw_values.vbs?"
listRawValueDone = ["Yes","Yes but I want to --replay lastly recorded macro-- by myself", "No"]
image="Yes.png"
RawValueDoneIn = eg.choicebox(questionRawValueDone, title1, listRawValueDone, preselect=1)
#RawValueDoneIn = eg.ynbox(questionRawValueDone)
#print("my answer is",RawValueDoneIn)
if RawValueDoneIn == listRawValueDone[2]:#"No":
	image = "No.gif"
	msg = "you'll get 7 seconds to clic on --create raw value-- under --Value--"
	choices = ["Yes","No","No opinion"]
	reply = eg.buttonbox(msg, image=image, choices=choices)	
	oShell.AppActivate('CANalyzer.Application')
	path2Vbs = r"O:\DEV_MS\01_MODELSHOP\20_RENAULT-NISSAN\02_AIVC1\01_PRODUCTION\03_TOOLS\Macros\Raw_values\Raw_values.vbs"
	os.system(path2Vbs)

elif RawValueDoneIn == listRawValueDone[0]:#"Yes":
	image = "Yes.gif"
	msg = "you'll get 5 seconds to clic on the diag --22 40 00--"
	choices = ["Yes","No","No opinion"]
	reply = eg.buttonbox(msg, image=image, choices=choices)	
	oShell.AppActivate('CANalyzer.Application')
	time.sleep(5)
	for i in range(13):
		oShell.sendkeys ("{ENTER}")
		oShell.SendKeys ("{DOWN}")
		time.sleep(0.15)
	for i in range(11):
		oShell.sendkeys ("{ENTER}")
		oShell.SendKeys ("{DOWN}")
		time.sleep(0.75)




checklist = 'test.xlsm'

ls_id=['Vehicle definition','AIVC Activation status', 'USB Definition', 'Ecall Zone', 'Car Variant', 'DataRecord', 'Vehicle Manufacturer Spare Part R', 'ConfigurationFileReferenceLink data',
 'VehicleManuECUHWNumber', 'System Supplier ECU SW', 'HWVariantID', 'TCU state','IMEI', 'CCID', 'Identificatio (IMSI)', 'eUICC EID','eUICC MSISDN', 'MQTT payload compression']

ls_pos=['E46', 'E47', 'E48' , 'E49', 'E50' , 'E51' , 'E53', 'E54', 'E55', 'E56', 'E57' , 'E58' , 'E59' , 'E60', 'E61', 'E62', 'E63', 'E64']

L = list(zip(ls_id, ls_pos))

data=[]

def core():
	with open("traceLog.txt", "r") as f:
		l = []
		for line in f : 
			line = line.strip()
			l = line.split('\t') 
			for elem in L :
				for k in l :
					if elem[0] in k :
						try : 
							print(elem[0] + ' =  '+l[l.index(k)+1] )
							L.remove(elem)
							data.append((elem[0], elem[1] ,l[l.index(k)+1]))
						except IndexError : 
							print(l[l.index(k)])

def fill_checklist(path, data_ls): 
	ww = O.load_workbook(filename=path)
	#ww = O.load_workbook("D:\test.xlsm")
	ws = ww.worksheets[2]
	for elem in data_ls: 
		ws[elem[1]] = elem[0]

if __name__ == "__main__" : 
	core()
	fill_checklist(checklist, data)


#App.Open "C:\CANWIN\DEMO_CN\CANSYSTEMDEMO\CANSYSTEMDEMO.CFG"
#Set Measurement = App.Measurement
#If App.Measurement.Running Then
#  App.Measurement.Stop
#End If
#Set Panels = App.Configuration.GeneralSetup.PanelSetup.Panels(0)
#Set Engine = Panels("Motor")
#Engine.Visible = True
#Set Pedal = App.Environment.GetVariable("EnvPedalPosition")
#Set Pedal.Value = 5

# Set App = CreateObject("CANalyzer.Application")
# Set Measurement = App.Measurement
# Wscript.ConnectObject Measurement, "Measurement_"

# Sub Measurement_OnInit()
#   Set TestFunction = App.CAPL.GetFunction("f")
# End Sub
# class ApplicationEvents(object):
#     def OnQuit(self):
#         print("quitting")

# app = win32com.client.DispatchWithEvents("CANalyzer.Application", ApplicationEvents)

#ReadDataByIdentifier_Process3410
#22 f3 ff

#https://github.com/weltond/AutomationWithCANalyzer/blob/master/canalyzer_test.py
#https://media.readthedocs.org/pdf/python-can/stable/python-can.pdf


# #run the can
# function1 = None
# canoe_app = None
# is_running = False

# class EventHandler:

# 	def OnInit(self):
# 		global canoe_app
# 		global function1

# 		function1 = canoe_app.CAPL.GetFunction('function1')

# 	def OnStart(self):
# 		global is_running
# 		is_running = True

# canoe_app = comclt.Dispatch('CANalyzer.Application')
# measurement = canoe_app.Measurement
# measurement_events = comclt.WithEvents(measurement, EventHandler)
# measurement.Start()
# # The following loop takes care of any pending events and, once, the Measurement
# # starts, it will call the CAPL function "function1" 10 times and then exit!
# count = 0

# if (is_running):
# 	function1.Call(count)
# 	count += 1

# pythoncom.PumpWaitingMessages()
# time.sleep(1)
# #def can():
# #	CANalyzer = comclt.Dispatch('CANalyzer.Application') 
# #	measurement = CANalyzer.Measurement
# 	#CANalyzer.Open("Configuration3.cfg") 
# 	#measurement.Start()
# 	#time.sleep(float(10)) # delays for XX seconds
# #	ref=CANalyzer.CAPL.GetFunction("function1")

# 	#measurement.Stop()
# 	#return 1
# #can()


# def send_one():
# 	#bus = can.interface.Bus(bustype='pcan', channel='PCAN_USBBUS1', bitrate=250000)
# 	#bus = can.interface.Bus(bustype='ixxat', channel=0, bitrate=250000)
# 	bus = can.interface.Bus(bustype='vector', app_name='CANalyzer', channel=0, bitrate=115200)

# 	msg = can.Message(arbitration_id=0xc0ffee,
# 					  data=[0, 25, 0, 1, 3, 1, 4, 1],
# 					  extended_id=True)
# 	try:
# 		bus.send(msg)
# 		print("Message sent on {}".format(bus.channel_info))
# 	except can.CanError:
# 		print("Message NOT sent")


# send_one()
